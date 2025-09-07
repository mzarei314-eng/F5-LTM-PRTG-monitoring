"""
PRTG F5 Monitoring Automation Script
------------------------------------
This script automates adding SNMP Custom Sensors in PRTG from an Excel file
using Selenium (Firefox + geckodriver).
"""

# imports...
from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
import openpyxl
import easygui
import pyperclip
from selenium.webdriver.firefox.options import Options

path=easygui.enterbox("Enter your path of excel file")
ex=openpyxl.load_workbook(fr"{path}")
sheet_choose=easygui.enterbox("Enter sheet name:")
sheet=ex[sheet_choose]



driver = webdriver.Firefox(executable_path=r'C:\Program Files\Mozilla Firefox\geckodriver.exe')
driver.get("http://192.0.2.10/")


easygui.msgbox("Enter admin user and password and continiue.")



Devices = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/header/nav/div/ul/li[2]/a'))
)
driver.execute_script("arguments[0].click();", Devices)

easygui.msgbox("select your device....")




add_sensor = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div/div/div/div/div[1]/div[2]/div[2]/a[1]'))
)
driver.execute_script("arguments[0].click();", add_sensor)





add_custom_sensor = WebDriverWait(driver, 10).until(
    EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[2]/div[1]/div[2]/div[1]/ul/li[3]/div'))
)
driver.execute_script("arguments[0].click();", add_custom_sensor)




row_select=int(easygui.enterbox("Enter your first row for adding."))
while str(sheet.cell(row=row_select,column=2).value )!= "None":
    oid = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.ID, 'oid_'))
    )
    driver.execute_script("arguments[0].click();", oid)

    driver.find_element(By.ID, 'name_').send_keys(Keys.CONTROL, "a")
    driver.find_element(By.ID, 'name_').send_keys(sheet.cell(row=row_select, column=1).value)
    driver.find_element(By.ID, 'oid_').send_keys(Keys.CONTROL, "a")
    driver.find_element(By.ID, 'oid_').send_keys(sheet.cell(row=row_select,column=2).value)
    time.sleep(1)
    element = driver.find_element(By.XPATH, '/html/body/div[1]/div/div/form/fieldset[3]/legend/span/label')
    driver.execute_script("arguments[0].click();", element)

    interval = driver.find_element(By.ID, 'interval_')
    driver.execute_script("arguments[0].click();", interval)

    time.sleep(2)
    driver.find_element(By.ID, 'interval_').send_keys(Keys.ARROW_UP,Keys.ENTER)
    time.sleep(2)
    driver.find_element(By.XPATH, '/html/body/div[1]/div/div/form/div/div/input[1]').click()
    try:
        add_sensor = WebDriverWait(driver, 60).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div/div/div/div/div[1]/div[2]/div[2]/a[1]'))
        )
        driver.execute_script("arguments[0].click();", add_sensor)

        add_custom_sensor = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[2]/div[1]/div[2]/div[1]/ul/li[3]/div'))
        )
        driver.execute_script("arguments[0].click();", add_custom_sensor)
        row_select += 1
    except Exception as e:
        #
        device=WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/header/div[3]/div[1]/ul/li[5]/a'))
        )
        driver.execute_script("arguments[0].click();", device)

        add_sensor = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div/div/div/div/div[1]/div[2]/div[2]/a[1]'))
        )
        driver.execute_script("arguments[0].click();", add_sensor)

        add_custom_sensor = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, '/html/body/div[1]/div/div/div[2]/div[1]/div[2]/div[1]/ul/li[3]/div'))
        )
        driver.execute_script("arguments[0].click();", add_custom_sensor)
        row_select += 1



easygui.msgbox("Done!!!")
